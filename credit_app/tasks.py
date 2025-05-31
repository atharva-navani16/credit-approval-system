from celery import shared_task
from django.conf import settings
import openpyxl
import os
from .models import Customer, Loan
from datetime import datetime
from decimal import Decimal

@shared_task
def ingest_customer_data():
    """Ingest customer data from Excel file"""
    file_path = os.path.join(settings.BASE_DIR, 'data', 'customer_data.xlsx')
    
    if not os.path.exists(file_path):
        return f"Customer data file not found at {file_path}"
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        customers_created = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None: 
                continue
                
            excel_customer_id, first_name, last_name, phone_number, monthly_salary, approved_limit, current_debt = row
            
            customer, created = Customer.objects.get_or_create(
                phone_number=phone_number,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'monthly_salary': Decimal(str(monthly_salary)),
                    'approved_limit': Decimal(str(approved_limit)),
                    'current_debt': Decimal(str(current_debt or 0)),
                    'age': 30 
                }
            )
            
            if created:
                customers_created += 1
        
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT setval(pg_get_serial_sequence('customers', 'customer_id'), COALESCE(MAX(customer_id), 1)) FROM customers;")
        
        return f"Successfully ingested {customers_created} customers"
        
    except Exception as e:
        return f"Error ingesting customer data: {str(e)}"

@shared_task
def ingest_loan_data():
    """Ingest loan data from Excel file"""
    file_path = os.path.join(settings.BASE_DIR, 'data', 'loan_data.xlsx')
    
    if not os.path.exists(file_path):
        return f"Loan data file not found at {file_path}"
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        loans_created = 0
        # Create a mapping of original customer IDs to new customer IDs
        customer_mapping = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
                
            excel_customer_id, loan_id, loan_amount, tenure, interest_rate, monthly_repayment, emis_paid_on_time, start_date, end_date = row
            
            try:
                if excel_customer_id not in customer_mapping:
                    customers = list(Customer.objects.all().order_by('customer_id'))
                    if excel_customer_id <= len(customers):
                        customer = customers[excel_customer_id - 1]  
                        customer_mapping[excel_customer_id] = customer
                    else:
                        continue  
                else:
                    customer = customer_mapping[excel_customer_id]
                
                if isinstance(start_date, str):
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                if isinstance(end_date, str):
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                
                loan = Loan.objects.create(
                    customer=customer,
                    loan_amount=Decimal(str(loan_amount)),
                    tenure=tenure,
                    interest_rate=Decimal(str(interest_rate)),
                    monthly_repayment=Decimal(str(monthly_repayment)),
                    emis_paid_on_time=emis_paid_on_time,
                    start_date=start_date,
                    end_date=end_date
                )
                
                loans_created += 1
                    
            except Exception as e:
                print(f"Error processing loan row: {e}")
                continue 
        
        return f"Successfully ingested {loans_created} loans"
        
    except Exception as e:
        return f"Error ingesting loan data: {str(e)}"